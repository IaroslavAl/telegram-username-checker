import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from transliterate import translit
from data_base.common_words import COMMON_WORDS
from data_base.names import get_names
from data_base.top_companies import top_companies
from data_base.crypto import crypto_channel_names
from data_base.channel_names import channel_names

def check_username_availability(username):
    url = f"https://fragment.com/?query={username}"

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        tm_value = None
        tm_element = soup.find('div', class_='tm-value') or soup.find('td', class_='tm-value')
        if tm_element:
            tm_value = tm_element.text.strip().lower()

        status = None
        status_element = soup.find('div', class_='tm-status-unavail') or \
                         soup.find('td', class_='wide-last-col') and soup.find('td', class_='wide-last-col').find('div',
                                                                                                                  class_='tm-value')
        if status_element:
            status = status_element.text.strip()

        is_unavailable = (tm_value == f"@{username.lower()}") and (status and 'unavailable' in status.lower())

        return {
            'username': username,
            'tm_value': tm_value,
            'status': status,
            'is_unavailable': is_unavailable,
            'error': None,
            'url': url,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    except Exception as e:
        return {
            'username': username,
            'tm_value': None,
            'status': None,
            'is_unavailable': False,
            'error': str(e),
            'url': url,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }


def create_excel_report(results, filename='fragment_report.xlsx'):
    wb = Workbook()

    # Основной лист с данными
    ws_main = wb.active
    ws_main.title = "Results"

    headers = ["Timestamp", "Username", "TM Value", "Status", "Unavailable", "Error", "URL"]
    ws_main.append(headers)

    # Стили для заголовков
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        cell = ws_main.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Стили для данных
    unavailable_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    available_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for result in results:
        row = [
            result['timestamp'],
            result['username'],
            result['tm_value'],
            result['status'],
            "Yes" if result['is_unavailable'] else "No",
            result['error'],
            result['url']
        ]
        ws_main.append(row)

        # Применяем цветовое форматирование
        last_row = ws_main.max_row
        if result['error']:
            for col in range(1, len(headers) + 1):
                ws_main.cell(row=last_row, column=col).fill = error_fill
        elif result['is_unavailable']:
            for col in range(1, len(headers) + 1):
                ws_main.cell(row=last_row, column=col).fill = unavailable_fill
        else:
            ws_main.cell(row=last_row, column=5).fill = available_fill

    # Лист с статистикой
    ws_stats = wb.create_sheet("Statistics")

    stats_headers = ["Metric", "Value"]
    ws_stats.append(stats_headers)

    total = len(results)
    unavailable = sum(1 for r in results if r['is_unavailable'])
    errors = sum(1 for r in results if r['error'])

    stats_data = [
        ["Total usernames checked", total],
        ["Unavailable usernames", unavailable],
        ["Available usernames", total - unavailable - errors],
        ["Errors", errors],
        ["Success rate", f"{(total - errors) / total * 100:.2f}%"]
    ]

    for data in stats_data:
        ws_stats.append(data)

    # Авто-ширина колонок
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)


def process_usernames(usernames, output_file='fragment_report.xlsx', delay=1):
    results = []

    for i, username in enumerate(usernames, 1):
        print(f"Processing {i}/{len(usernames)}: {username}")

        result = check_username_availability(username)
        results.append(result)

        if result['error']:
            print(f"  Error: {result['error']}")
        else:
            print(
                f"  Result: tm-value='{result['tm_value']}', status='{result['status']}', unavailable={result['is_unavailable']}")

        if i < len(usernames):
            time.sleep(delay)

    create_excel_report(results, output_file)
    print(f"\nReport saved to {output_file}")


if __name__ == "__main__":
    # Ваш список username'ов
    usernames_to_check = channel_names

    process_usernames(
        usernames=usernames_to_check,
        output_file='fragment_usernames_report.xlsx',
        delay=2  # Задержка между запросами (в секундах)
    )